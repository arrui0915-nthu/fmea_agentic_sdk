"""Convert one FMEA Excel workbook into one row-delimited Markdown file."""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet


STANDARD_FIELDS = [
    "process",
    "functional_requirement",
    "potential_failure_mode",
    "potential_failure_effect",
    "severity_before",
    "potential_causes",
    "occurrence_before",
    "current_process_controls",
    "detection_before",
    "rpn_before",
    "recommended_actions",
    "severity_after",
    "occurrence_after",
    "detection_after",
    "rpn_after",
    "owner_date",
]


@dataclass(frozen=True)
class ColumnMapping:
    """Header aliases and the occurrence to use for duplicated headers."""

    aliases: tuple[str, ...]
    occurrence: int = 1


# This is the single source of truth for mapping workbook headers to standard fields.
# ``occurrence`` distinguishes the before/after columns when their labels are identical.
COLUMN_MAPPINGS: dict[str, ColumnMapping] = {
    "process": ColumnMapping(
        ("設計製程項目", "製程項目", "process design", "process")
    ),
    "functional_requirement": ColumnMapping(
        ("功能要求", "functional requirement")
    ),
    "potential_failure_mode": ColumnMapping(
        ("潛在失效模式", "失效模式", "potential failure mode", "potential failure model")
    ),
    "potential_failure_effect": ColumnMapping(
        ("失效可能後果", "潛在失效後果", "failure effect", "failure causes potential serious results")
    ),
    "severity_before": ColumnMapping(("嚴重度", "severity"), occurrence=1),
    "potential_causes": ColumnMapping(
        ("失效可能原因", "潛在失效原因", "potential causes", "potential reasons")
    ),
    "occurrence_before": ColumnMapping(("發生度", "occurrence"), occurrence=1),
    "current_process_controls": ColumnMapping(
        ("現有控制方法", "現行製程管制", "current process controls", "current process inspection")
    ),
    "detection_before": ColumnMapping(("偵測度", "探測度", "detection"), occurrence=1),
    "rpn_before": ColumnMapping(("風險等級", "風險優先數", "rpn"), occurrence=1),
    "recommended_actions": ColumnMapping(("建議措施", "recommended actions")),
    "severity_after": ColumnMapping(("嚴重度", "severity"), occurrence=2),
    "occurrence_after": ColumnMapping(("發生度", "occurrence"), occurrence=2),
    "detection_after": ColumnMapping(("偵測度", "探測度", "detection"), occurrence=2),
    "rpn_after": ColumnMapping(("風險等級", "風險優先數", "rpn"), occurrence=2),
    "owner_date": ColumnMapping(("負責人日期", "負責人/日期", "owner date", "responsible person date")),
}

HIERARCHICAL_FIELDS = ("process",)
MIN_HEADER_MATCHES = 6


@dataclass(frozen=True)
class FmeaRow:
    source_sheet: str
    source_excel_row: int
    values: dict[str, str]


class FmeaExcelError(ValueError):
    """Raised when a workbook cannot be interpreted as an FMEA workbook."""


def _normalise_header(value: Any) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"[\s\u3000:：()（）/\\_\-]+", "", text).casefold()


def _header_matches(value: Any, aliases: tuple[str, ...]) -> bool:
    header = _normalise_header(value)
    return bool(header) and any(
        _normalise_header(alias) in header for alias in aliases
    )


def _column_map_for_row(ws: Worksheet, header_row: int) -> dict[str, int]:
    headers = {
        column: ws.cell(row=header_row, column=column).value
        for column in range(1, ws.max_column + 1)
    }
    result: dict[str, int] = {}
    for field, mapping in COLUMN_MAPPINGS.items():
        candidates = [
            column
            for column, value in headers.items()
            if _header_matches(value, mapping.aliases)
        ]
        if len(candidates) >= mapping.occurrence:
            result[field] = candidates[mapping.occurrence - 1]
    return result


def detect_fmea_sheet_and_header(workbook: Any) -> tuple[Worksheet, int, dict[str, int]]:
    """Return the worksheet/header row with the strongest standard-field match."""

    best: tuple[int, Worksheet, int, dict[str, int]] | None = None
    for ws in workbook.worksheets:
        for row_number in range(1, min(ws.max_row, 50) + 1):
            column_map = _column_map_for_row(ws, row_number)
            score = len(column_map)
            if best is None or score > best[0]:
                best = (score, ws, row_number, column_map)

    if best is None or best[0] < MIN_HEADER_MATCHES:
        raise FmeaExcelError(
            "找不到 FMEA 欄位標題列；至少需要辨識 "
            f"{MIN_HEADER_MATCHES} 個標準欄位。"
        )

    _, worksheet, header_row, column_map = best
    required = {"process", "potential_failure_mode"}
    missing = sorted(required - column_map.keys())
    if missing:
        raise FmeaExcelError(f"FMEA 標題缺少必要欄位: {', '.join(missing)}")
    return worksheet, header_row, column_map


def _merged_cell_values(ws: Worksheet) -> dict[tuple[int, int], Any]:
    values: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        anchor = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                values[(row, column)] = anchor
    return values


def _clean_value(value: Any) -> str:
    if value is None or isinstance(value, MergedCell):
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip().replace("\r\n", "\n").replace("\r", "\n")


def read_fmea_rows(excel_path: Path) -> list[FmeaRow]:
    """Read and standardise all non-empty FMEA rows from a workbook."""

    excel_path = Path(excel_path)
    if not excel_path.is_file():
        raise FileNotFoundError(f"找不到 Excel: {excel_path}")

    workbook = load_workbook(excel_path, data_only=True, read_only=False)
    try:
        ws, header_row, column_map = detect_fmea_sheet_and_header(workbook)
        merged_values = _merged_cell_values(ws)
        forward_values = {field: "" for field in HIERARCHICAL_FIELDS}
        rows: list[FmeaRow] = []

        for row_number in range(header_row + 1, ws.max_row + 1):
            values: dict[str, str] = {}
            for field in STANDARD_FIELDS:
                column = column_map.get(field)
                raw_value = None
                if column is not None:
                    raw_value = merged_values.get(
                        (row_number, column), ws.cell(row_number, column).value
                    )
                values[field] = _clean_value(raw_value)

            # Check before forward fill so styled trailing rows do not become data rows.
            if not any(values.values()):
                continue

            for field in HIERARCHICAL_FIELDS:
                if values[field]:
                    forward_values[field] = values[field]
                else:
                    values[field] = forward_values[field]

            rows.append(
                FmeaRow(
                    source_sheet=ws.title,
                    source_excel_row=row_number,
                    values=values,
                )
            )
        return rows
    finally:
        workbook.close()


def _process_code(excel_path: Path) -> str:
    stem = re.sub(r"(?i)_?FMEA$", "", excel_path.stem)
    code = re.sub(r"[^A-Za-z0-9]+", "-", stem).strip("-").upper()
    if not code:
        raise FmeaExcelError(f"無法從檔名取得製程代碼: {excel_path.name}")
    return code


def _markdown_value(value: str) -> str:
    # Keep every cell on one Markdown list item while retaining embedded line breaks.
    return value.replace("\n", "<br>")


def render_fmea_markdown(excel_path: Path, rows: list[FmeaRow]) -> str:
    code = _process_code(excel_path)
    blocks: list[str] = []
    for sequence, row in enumerate(rows, start=1):
        document_id = f"{code}-{sequence:04d}"
        lines = [
            f"<!-- FMEA_ROW_START id={document_id} -->",
            "",
            f"## {document_id}",
            "",
            f"- source_excel: {excel_path.name}",
            f"- source_sheet: {_markdown_value(row.source_sheet)}",
            f"- source_excel_row: {row.source_excel_row}",
        ]
        lines.extend(
            f"- {field}: {_markdown_value(row.values[field])}"
            for field in STANDARD_FIELDS
        )
        lines.extend(("", "<!-- FMEA_ROW_END -->"))
        blocks.append("\n".join(lines))
    return "\n\n".join(blocks) + ("\n" if blocks else "")


def convert_excel_to_markdown(excel_path: Path, output_dir: Path) -> Path:
    """Convert one Excel workbook to exactly one Markdown output file."""

    excel_path = Path(excel_path)
    output_dir = Path(output_dir)
    rows = read_fmea_rows(excel_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{excel_path.stem}.md"
    output_path.write_text(render_fmea_markdown(excel_path, rows), encoding="utf-8")
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "excel_path",
        nargs="?",
        type=Path,
        default=Path("data/input/PVD_FMEA.xlsx"),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("data/markdown"),
    )
    args = parser.parse_args()
    output_path = convert_excel_to_markdown(args.excel_path, args.output_dir)
    print(output_path)


if __name__ == "__main__":
    main()
