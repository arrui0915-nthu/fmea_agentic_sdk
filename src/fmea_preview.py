"""Agentic SDK workflow for generating read-only FMEA previews from chat logs."""

from __future__ import annotations

import json
from collections.abc import Callable, Mapping, Sequence
from io import BytesIO
from typing import Any

import faiss
import numpy as np
from agentic_sdk import (
    ContextEntry,
    ContextEntryType,
    ModuleOutput,
    Workflow,
    WorkflowAborted,
    WorkflowState,
)
from agentic_sdk.llm import (
    OpenAIChatResponse,
    chat_stream_json,
    require_model,
    resolve_openai_client,
)
from openpyxl import Workbook, load_workbook

from src.config import Settings, load_settings
from src.excel_to_markdown import STANDARD_FIELDS
from src.faiss_knowledge_base import FmeaFaissKnowledgeBase


PREVIEW_SIMILARITY_THRESHOLD = 0.85
BUILD_FMEA_PREVIEW_TOOL_NAME = "build_fmea_preview"

NUMERIC_FIELDS = {
    "severity_before",
    "occurrence_before",
    "detection_before",
    "rpn_before",
    "severity_after",
    "occurrence_after",
    "detection_after",
    "rpn_after",
}

BUILD_FMEA_PREVIEW_TOOL: dict[str, Any] = {
    "type": "function",
    "function": {
        "name": BUILD_FMEA_PREVIEW_TOOL_NAME,
        "description": "將已通過相似度去重的候選資料正規化為 FMEA preview Excel。",
        "parameters": {
            "type": "object",
            "properties": {},
            "additionalProperties": False,
        },
    },
}

PREVIEW_EVENTS_SCHEMA = {
    "perceive": {"label": "整理聊天紀錄", "fields": ["candidate_rows"]},
    "plan": {"label": "規劃去重流程", "fields": []},
    "retrieve": {"label": "比對既有 FMEA", "fields": []},
    "action": {"label": "產生 FMEA 表格", "fields": []},
    "reflect": {"label": "驗證 Preview", "fields": []},
}

_ROW_TEMPLATE_JSON = json.dumps(
    {field: None for field in STANDARD_FIELDS}, ensure_ascii=False
)

PERCEIVE_SYSTEM_PROMPT = f"""你是 FMEA 資料整理器。請將聊天紀錄整理成候選 FMEA rows。

只回傳 JSON object，格式為：
{{"candidate_rows": [{_ROW_TEMPLATE_JSON}]}}

規則：
1. 每個不同的失效事件建立一個 row，同一事件在聊天中重複出現時合併成一個 row。
2. 每個 row 必須剛好包含指定的 {len(STANDARD_FIELDS)} 個欄位，不可增加欄位。
3. 可以將聊天中的內容整理成簡潔欄位文字，但不可加入聊天沒有依據的事實。
4. 不可自行猜測 severity、occurrence、detection、RPN、負責人或日期。
5. 資訊不足的欄位使用 JSON null，不要使用空字串、"NULL" 或其他占位文字。
6. 若聊天中沒有任何具體失效、風險、原因、效應或控制資訊，回傳空陣列。
7. 保持原本聊天使用的語言。
"""


ChatJsonRunner = Callable[..., OpenAIChatResponse]


def decode_chat_log(content: bytes) -> str:
    """Decode an uploaded chat log as UTF-8, accepting an optional BOM."""

    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError(
            "檔案不是有效的 UTF-8 文字，請轉成 UTF-8 後再上傳。"
        ) from exc


def _is_null(value: object) -> bool:
    return value is None or (
        isinstance(value, str)
        and value.strip().casefold() in {"", "null", "none", "n/a", "na"}
    )


def _normalise_number(value: object) -> int | float | None:
    if _is_null(value) or isinstance(value, bool):
        return None
    try:
        number = float(str(value).strip())
    except (TypeError, ValueError):
        return None
    if not np.isfinite(number):
        return None
    return int(number) if number.is_integer() else number


def normalise_fmea_row(raw_row: Mapping[str, Any]) -> dict[str, Any]:
    """Return one row with exactly the canonical fields and nullable values."""

    row: dict[str, Any] = {}
    for field in STANDARD_FIELDS:
        value = raw_row.get(field)
        if field in NUMERIC_FIELDS:
            row[field] = _normalise_number(value)
        elif _is_null(value):
            row[field] = None
        else:
            row[field] = str(value).strip()

    _fill_rpn(row, suffix="before")
    _fill_rpn(row, suffix="after")
    return row


def _fill_rpn(row: dict[str, Any], *, suffix: str) -> None:
    rpn_field = f"rpn_{suffix}"
    if row[rpn_field] is not None:
        return
    ratings = [
        row[f"severity_{suffix}"],
        row[f"occurrence_{suffix}"],
        row[f"detection_{suffix}"],
    ]
    if all(isinstance(value, (int, float)) for value in ratings):
        product = float(ratings[0]) * float(ratings[1]) * float(ratings[2])
        row[rpn_field] = int(product) if product.is_integer() else product


def normalise_fmea_rows(raw_rows: object) -> list[dict[str, Any]]:
    """Normalise candidate rows and discard structurally empty model output."""

    if not isinstance(raw_rows, list):
        return []
    rows: list[dict[str, Any]] = []
    seen: set[tuple[Any, ...]] = set()
    for raw_row in raw_rows:
        if not isinstance(raw_row, Mapping):
            continue
        row = normalise_fmea_row(raw_row)
        signature = tuple(row[field] for field in STANDARD_FIELDS)
        if not any(value is not None for value in signature) or signature in seen:
            continue
        seen.add(signature)
        rows.append(row)
    return rows


def display_fmea_rows(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    """Convert nullable rows to the visible/downloaded NULL representation."""

    return [
        {
            field: "NULL" if row.get(field) is None else row.get(field)
            for field in STANDARD_FIELDS
        }
        for row in rows
    ]


def build_fmea_workbook(rows: Sequence[Mapping[str, Any]]) -> bytes | None:
    """Build an in-memory workbook with the canonical FMEA header order."""

    if not rows:
        return None
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "FMEA Preview"
    worksheet.append(list(STANDARD_FIELDS))
    for row in display_fmea_rows(rows):
        worksheet.append([row[field] for field in STANDARD_FIELDS])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


class ConversationFmeaPerceive:
    """Extract candidate rows from an uploaded or pasted conversation."""

    name = "perceive"
    gen_ai_system = "openai_compatible"

    def __init__(
        self,
        *,
        api_key: str | None = None,
        base_url: str | None = None,
        model: str | None = None,
        client: Any | None = None,
        chat_runner: ChatJsonRunner = chat_stream_json,
    ) -> None:
        self._model = require_model(model, self.__class__.__name__)
        self._client = client or resolve_openai_client(
            self.__class__.__name__, api_key=api_key, base_url=base_url
        )
        self._chat_runner = chat_runner

    @property
    def gen_ai_request_model(self) -> str:
        return self._model

    def __call__(self, state: WorkflowState) -> ModuleOutput:
        conversation = state.latest_user_message().strip()
        if not conversation:
            rows: list[dict[str, Any]] = []
            usage = None
        else:
            try:
                response = self._chat_runner(
                    self._client,
                    model=self._model,
                    system=PERCEIVE_SYSTEM_PROMPT,
                    user=conversation,
                    on_delta=lambda content: state.emit_token_delta(
                        self.name,
                        content,
                        metadata={"model": self._model, "structured": True},
                    ),
                    structured_fields=state.structured_fields_for(self.name),
                    on_field=lambda field, value: state.emit_structured_field(
                        self.name,
                        field,
                        value,
                        metadata={"model": self._model, "structured": True},
                    ),
                )
                parsed = response.as_json()
                rows = normalise_fmea_rows(parsed.get("candidate_rows"))
                usage = {
                    "model": response.model or self._model,
                    "input_tokens": response.input_tokens,
                    "output_tokens": response.output_tokens,
                }
            except Exception as exc:
                message = "無法從聊天紀錄整理 FMEA 候選資料。"
                state.last_workflow_error = {"stage": self.name, "message": message}
                raise WorkflowAborted(message) from exc

        return ModuleOutput(
            next_module="plan",
            payload={"candidate_rows": rows, "_llm_usage": usage},
            context_updates=[
                ContextEntry(
                    type=ContextEntryType.PERCEIVED,
                    content=f"candidate_rows={len(rows)}",
                    metadata={"candidate_count": len(rows)},
                )
            ],
        )


class FmeaPreviewPlan:
    """Deterministically route extracted rows through the duplicate check."""

    name = "plan"

    def __call__(self, state: WorkflowState) -> ModuleOutput:
        candidate_count = len(state.lookup("candidate_rows") or [])
        return ModuleOutput(
            next_module="retrieve",
            payload={"preview_plan": "compare_existing_fmea"},
            context_updates=[
                ContextEntry(
                    type=ContextEntryType.PLAN_DECISION,
                    content="next=retrieve",
                    metadata={"candidate_count": candidate_count},
                )
            ],
        )


def _row_embedding_text(row: Mapping[str, Any]) -> str:
    return "\n".join(
        f"- {field}: {'' if row.get(field) is None else row.get(field)}"
        for field in STANDARD_FIELDS
    )


class CandidateSimilarityRetrieve:
    """Read existing FAISS indexes and reject candidates over a cosine threshold."""

    name = "retrieve"

    def __init__(
        self,
        knowledge_bases: Mapping[str, FmeaFaissKnowledgeBase],
        *,
        embedding_client: Any,
        embedding_model: str,
        threshold: float = PREVIEW_SIMILARITY_THRESHOLD,
    ) -> None:
        if not -1.0 <= float(threshold) <= 1.0:
            raise ValueError("similarity threshold 必須介於 -1 與 1")
        self.knowledge_bases = {
            str(code).upper(): kb for code, kb in knowledge_bases.items()
        }
        self.embedding_client = embedding_client
        self.embedding_model = embedding_model
        self.threshold = float(threshold)

    def __call__(self, state: WorkflowState) -> ModuleOutput:
        candidates = normalise_fmea_rows(state.lookup("candidate_rows") or [])
        accepted: list[dict[str, Any]] = []
        duplicates: list[dict[str, Any]] = []

        if candidates:
            vectors = self._embed_rows(candidates)
            for candidate_index, (row, vector) in enumerate(
                zip(candidates, vectors, strict=True), start=1
            ):
                best_match = self._best_match(vector)
                if best_match is not None and best_match["similarity"] >= self.threshold:
                    duplicates.append(
                        {
                            "candidate_index": candidate_index,
                            "row": row,
                            **best_match,
                        }
                    )
                else:
                    accepted.append(row)

        summary = (
            f"candidate_count={len(candidates)} accepted_count={len(accepted)} "
            f"duplicate_count={len(duplicates)} threshold={self.threshold:.2f}"
        )
        return ModuleOutput(
            next_module="action",
            payload={
                "candidate_rows": candidates,
                "accepted_rows": accepted,
                "duplicate_rows": duplicates,
                "similarity_threshold": self.threshold,
            },
            context_updates=[
                ContextEntry(
                    type=ContextEntryType.RETRIEVED,
                    content=summary,
                    metadata={
                        "candidate_count": len(candidates),
                        "accepted_count": len(accepted),
                        "duplicate_count": len(duplicates),
                        "hit_count": len(duplicates),
                        "threshold": self.threshold,
                    },
                )
            ],
        )

    def _embed_rows(self, rows: Sequence[Mapping[str, Any]]) -> np.ndarray:
        response = self.embedding_client.embeddings.create(
            input=[_row_embedding_text(row) for row in rows],
            model=self.embedding_model,
        )
        response_data = sorted(
            response.data, key=lambda item: int(getattr(item, "index", 0))
        )
        vectors = np.asarray(
            [item.embedding for item in response_data], dtype="float32"
        )
        if vectors.ndim != 2 or vectors.shape[0] != len(rows) or vectors.shape[1] == 0:
            raise ValueError("候選 FMEA embedding 數量或維度不正確")
        faiss.normalize_L2(vectors)
        return vectors

    def _best_match(self, vector: np.ndarray) -> dict[str, Any] | None:
        best: dict[str, Any] | None = None
        query = np.asarray([vector], dtype="float32")
        for process_code, knowledge_base in self.knowledge_bases.items():
            index = knowledge_base.index
            if index is None or not knowledge_base.documents or int(index.ntotal) == 0:
                continue
            if int(index.d) != int(query.shape[1]):
                raise ValueError(
                    f"{process_code} index 維度與候選 embedding 不一致"
                )
            scores, indices = index.search(query, 1)
            document_index = int(indices[0][0])
            if document_index < 0:
                continue
            similarity = float(scores[0][0])
            document = knowledge_base.documents[document_index]
            match = {
                "matched_process": process_code,
                "matched_document_id": document.document_id,
                "similarity": similarity,
            }
            if best is None or similarity > float(best["similarity"]):
                best = match
        return best


class FmeaPreviewToolDispatcher:
    """Allow-list and execute the deterministic preview builder tool."""

    def execute(
        self,
        name: str,
        arguments: Mapping[str, Any],
        *,
        rows: Sequence[Mapping[str, Any]],
    ) -> dict[str, Any]:
        if name != BUILD_FMEA_PREVIEW_TOOL_NAME:
            raise ValueError(f"不允許的工具：{name}")
        if arguments:
            raise ValueError("build_fmea_preview 不接受參數")
        preview_rows = [normalise_fmea_row(row) for row in rows]
        return {
            "preview_rows": preview_rows,
            "preview_xlsx_bytes": build_fmea_workbook(preview_rows),
        }


class FmeaPreviewToolAction:
    """Action stage that calls only the allow-listed workbook generation tool."""

    name = "action"

    def __init__(self, dispatcher: FmeaPreviewToolDispatcher | None = None) -> None:
        self.dispatcher = dispatcher or FmeaPreviewToolDispatcher()

    def __call__(self, state: WorkflowState) -> ModuleOutput:
        accepted = state.lookup("accepted_rows") or []
        try:
            if accepted:
                tool_result = self.dispatcher.execute(
                    BUILD_FMEA_PREVIEW_TOOL_NAME,
                    {},
                    rows=accepted,
                )
                message = f"已產生 {len(tool_result['preview_rows'])} 筆 FMEA Preview。"
                tool_calls = [BUILD_FMEA_PREVIEW_TOOL_NAME]
            else:
                tool_result = {
                    "preview_rows": [],
                    "preview_xlsx_bytes": None,
                }
                message = "沒有可加入的新 FMEA rows。"
                tool_calls = []
        except Exception as exc:
            state.last_action_error = {
                "type": type(exc).__name__,
                "message": str(exc),
            }
            return ModuleOutput(
                next_module=None,
                payload={
                    "preview_rows": [],
                    "preview_xlsx_bytes": None,
                },
                context_updates=[
                    ContextEntry(
                        type=ContextEntryType.ACTION_RESULT,
                        content=f"error:{type(exc).__name__}",
                        metadata={"ok": False, "error": str(exc)},
                    )
                ],
            )

        state.last_action_error = None
        state.last_action_result = {
            "content": message,
            "model": "deterministic-fmea-preview-tool",
            "tool_calls": tool_calls,
        }
        return ModuleOutput(
            next_module=None,
            payload={
                "latest_final_message": message,
                **tool_result,
            },
            context_updates=[
                ContextEntry(
                    type=ContextEntryType.ACTION_RESULT,
                    content=message,
                    metadata={"ok": True, "tool_names": tool_calls},
                )
            ],
        )


class FmeaPreviewReflect:
    """Validate that preview rows and workbook agree without requiring hits."""

    name = "reflect"

    def __call__(self, state: WorkflowState) -> ModuleOutput:
        error = state.last_action_error
        reason = str(error.get("message")) if error else self._validation_error(state)
        verdict = "fail" if reason else "pass"
        return ModuleOutput(
            next_module=None,
            payload={"reflect_verdict": verdict},
            context_updates=[
                ContextEntry(
                    type=ContextEntryType.REFLECTION,
                    content=f"verdict={verdict}",
                    metadata={
                        "verdict": verdict,
                        "reason": reason or "preview schema and workbook are valid",
                    },
                )
            ],
        )

    @staticmethod
    def _validation_error(state: WorkflowState) -> str | None:
        accepted = state.lookup("accepted_rows") or []
        preview = state.lookup("preview_rows") or []
        workbook_bytes = state.lookup("preview_xlsx_bytes")
        if len(preview) != len(accepted):
            return "preview row count 與 accepted row count 不一致"
        if any(list(row.keys()) != list(STANDARD_FIELDS) for row in preview):
            return "preview schema 不正確"
        if not preview:
            return None if workbook_bytes is None else "空 preview 不應產生 workbook"
        if not isinstance(workbook_bytes, bytes) or not workbook_bytes:
            return "preview workbook 不存在"
        try:
            workbook = load_workbook(BytesIO(workbook_bytes), read_only=True)
            worksheet = workbook.active
            headers = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
            row_count = worksheet.max_row - 1
            workbook.close()
        except Exception as exc:
            return f"preview workbook 無法讀取：{exc}"
        if headers != list(STANDARD_FIELDS):
            return "preview workbook header 不正確"
        if row_count != len(preview):
            return "preview workbook row count 不正確"
        return None


def build_preview_workflow(
    knowledge_bases: Mapping[str, FmeaFaissKnowledgeBase],
    *,
    selected_processes: Sequence[str] | None = None,
    threshold: float = PREVIEW_SIMILARITY_THRESHOLD,
    settings: Settings | None = None,
    embedding_client: Any | None = None,
    perceive: Any | None = None,
) -> Workflow:
    """Build the isolated preview workflow without mutating any knowledge base."""

    settings = settings or load_settings(require_chat=True, require_embedding=True)
    available = {str(code).upper(): kb for code, kb in knowledge_bases.items()}
    requested = (
        list(available)
        if selected_processes is None
        else [str(code).strip().upper() for code in selected_processes]
    )
    selected = {
        code: available[code]
        for code in dict.fromkeys(requested)
        if code in available
    }
    if not selected:
        raise ValueError("至少需要選擇一個可用的 FMEA 製程")

    resolved_embedding_client = embedding_client or next(
        iter(selected.values())
    ).embedding_client
    resolved_embedding_model = str(settings.embedding_model)
    perceive_module = perceive or ConversationFmeaPerceive(
        api_key=settings.chat_api_key,
        base_url=settings.chat_base_url,
        model=settings.chat_model,
    )
    return Workflow(
        workflow_name="聊天紀錄 FMEA Preview",
        description="從聊天紀錄產生去重後、可下載但不發布的 FMEA Preview。",
        perceive=perceive_module,
        plan=FmeaPreviewPlan(),
        retrieve=CandidateSimilarityRetrieve(
            selected,
            embedding_client=resolved_embedding_client,
            embedding_model=resolved_embedding_model,
            threshold=threshold,
        ),
        action=FmeaPreviewToolAction(),
        reflect=FmeaPreviewReflect(),
        events_schema=PREVIEW_EVENTS_SCHEMA,
    )


def duplicate_rows_for_display(
    duplicates: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    """Flatten internal duplicate metadata for Streamlit display."""

    result: list[dict[str, Any]] = []
    for duplicate in duplicates:
        row = duplicate.get("row")
        if not isinstance(row, Mapping):
            row = {}
        result.append(
            {
                "candidate_index": duplicate.get("candidate_index"),
                "process": row.get("process") or "NULL",
                "potential_failure_mode": row.get("potential_failure_mode") or "NULL",
                "matched_process": duplicate.get("matched_process"),
                "matched_document_id": duplicate.get("matched_document_id"),
                "similarity": round(float(duplicate.get("similarity") or 0.0), 4),
            }
        )
    return result
