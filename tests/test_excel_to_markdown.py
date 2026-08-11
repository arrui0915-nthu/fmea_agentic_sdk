from pathlib import Path

import pytest
from openpyxl import Workbook

from src.excel_to_markdown import (
    OPTIONAL_FIELDS,
    SOURCE_FIELDS,
    STANDARD_FIELDS,
    convert_excel_to_markdown,
    read_fmea_rows,
)
from src.my_splitter import split_fmea_markdown


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PVD_EXCEL = PROJECT_ROOT / "data" / "input" / "PVD_FMEA.xlsx"


def _create_fmea_workbook(
    tmp_path: Path,
    *,
    machine_action_header: str | None = None,
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "fmea_pvd"
    headers = [
        "process",
        "functional requirement",
        "potential failure mode",
        "failure effect",
        "severity",
        "potential causes",
        "occurrence",
        "current process controls",
        "detection",
        "rpn",
        "recommended actions",
        "severity",
        "occurrence",
        "detection",
        "rpn",
        "owner date",
    ]
    values: list[object] = [
        "PVD",
        "stable coating",
        "coating too thin",
        "low durability",
        8,
        "low chamber pressure",
        4,
        "inspect pressure",
        1,
        32,
        "adjust setpoints",
        4,
        2,
        1,
        8,
        "owner / 2026-08-11",
    ]
    if machine_action_header is not None:
        headers.append(machine_action_header)
        values.append(
            '{"machine_id":"PVD-DEMO-01","setpoints":'
            '{"button_1":10,"button_2":20,"button_3":30}}'
        )
    worksheet.append(headers)
    worksheet.append(values)
    excel_path = tmp_path / "PVD_FMEA.xlsx"
    workbook.save(excel_path)
    workbook.close()
    return excel_path


def test_one_excel_produces_one_markdown(tmp_path: Path) -> None:
    output_path = convert_excel_to_markdown(PVD_EXCEL, tmp_path)

    assert output_path == tmp_path / "PVD_FMEA.md"
    assert list(tmp_path.glob("*.md")) == [output_path]


def test_valid_excel_rows_equal_documents_and_expected_count(tmp_path: Path) -> None:
    rows = read_fmea_rows(PVD_EXCEL)
    output_path = convert_excel_to_markdown(PVD_EXCEL, tmp_path)
    documents = split_fmea_markdown(output_path)

    assert len(rows) == len(documents) == 10


def test_process_is_forward_filled() -> None:
    rows = read_fmea_rows(PVD_EXCEL)

    assert [row.source_excel_row for row in rows] == list(range(3, 13))
    assert all(row.values["process"] == "1. 晶圓進料" for row in rows[:4])
    assert all(row.values["process"] == "2. PVD製程" for row in rows[4:])


def test_every_document_has_original_excel_row_and_unique_id(tmp_path: Path) -> None:
    output_path = convert_excel_to_markdown(PVD_EXCEL, tmp_path)
    documents = split_fmea_markdown(output_path)

    assert [document.metadata["source_excel_row"] for document in documents] == list(
        range(3, 13)
    )
    ids = [document.document_id for document in documents]
    assert ids == [f"PVD-{number:04d}" for number in range(1, 11)]
    assert len(ids) == len(set(ids))


def test_empty_cells_never_render_as_none(tmp_path: Path) -> None:
    output_path = convert_excel_to_markdown(PVD_EXCEL, tmp_path)

    assert "None" not in output_path.read_text(encoding="utf-8")


def test_standard_fields_schema_remains_unchanged() -> None:
    assert STANDARD_FIELDS == [
        "process",
        "functional_requirement",
        "potential_failure_mode",
        "potential_failure_effect",
        "severity_before",
        "potential_causes",
        "occurrence_before",
        "current_process_controls",
        "detection_before",
        "rpn_before",
        "recommended_actions",
        "severity_after",
        "occurrence_after",
        "detection_after",
        "rpn_after",
        "owner_date",
    ]
    assert OPTIONAL_FIELDS == ["machine_action"]
    assert SOURCE_FIELDS == [*STANDARD_FIELDS, "machine_action"]


@pytest.mark.parametrize(
    "header",
    ["machine_action", "machine action", "機台動作"],
)
def test_optional_machine_action_is_read_and_rendered(
    tmp_path: Path,
    header: str,
) -> None:
    excel_path = _create_fmea_workbook(
        tmp_path,
        machine_action_header=header,
    )

    rows = read_fmea_rows(excel_path)
    output_path = convert_excel_to_markdown(excel_path, tmp_path / "markdown")

    expected = (
        '{"machine_id":"PVD-DEMO-01","setpoints":'
        '{"button_1":10,"button_2":20,"button_3":30}}'
    )
    assert rows[0].values["machine_action"] == expected
    assert f"- machine_action: {expected}" in output_path.read_text(encoding="utf-8")
    assert split_fmea_markdown(output_path)[0].metadata["machine_action"] == expected


def test_missing_optional_machine_action_is_empty_and_rendered(
    tmp_path: Path,
) -> None:
    excel_path = _create_fmea_workbook(tmp_path)

    rows = read_fmea_rows(excel_path)
    output_path = convert_excel_to_markdown(excel_path, tmp_path / "markdown")

    assert rows[0].values["machine_action"] == ""
    assert "- machine_action: \n" in output_path.read_text(encoding="utf-8")


def test_detects_fmea_sheet_and_header_row(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "說明"
    workbook.active["A1"] = "這不是 FMEA 資料"
    worksheet = workbook.create_sheet("target_fmea")
    worksheet.append(["報表名稱"])
    worksheet.append([])
    worksheet.append(
        [
            "製程項目",
            "功能要求",
            "潛在失效模式",
            "失效可能後果",
            "嚴重度",
            "失效可能原因",
            "發生度",
            "現有控制方法",
            "偵測度",
            "風險等級(RPN)",
            "建議措施",
            "嚴重度",
            "發生度",
            "偵測度",
            "風險等級(RPN)",
            "負責人/日期",
        ]
    )
    worksheet.append(
        ["PVD", "需求", "失效", "後果", 8, "原因", 4, "控制", 1, 32]
    )
    excel_path = tmp_path / "PVD_FMEA.xlsx"
    workbook.save(excel_path)

    rows = read_fmea_rows(excel_path)

    assert len(rows) == 1
    assert rows[0].source_sheet == "target_fmea"
    assert rows[0].source_excel_row == 4


def test_repeated_conversion_is_deterministic(tmp_path: Path) -> None:
    output_path = convert_excel_to_markdown(PVD_EXCEL, tmp_path)
    first_result = output_path.read_bytes()

    repeated_path = convert_excel_to_markdown(PVD_EXCEL, tmp_path)

    assert repeated_path == output_path
    assert repeated_path.read_bytes() == first_result
