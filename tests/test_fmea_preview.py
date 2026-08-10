from __future__ import annotations

from io import BytesIO
from types import SimpleNamespace
from typing import Any

import faiss
import numpy as np
import pytest
from agentic_sdk import Workflow, WorkflowAborted, WorkflowState
from agentic_sdk.llm import OpenAIChatResponse
from openpyxl import load_workbook

from src.excel_to_markdown import STANDARD_FIELDS
from src.fmea_preview import (
    BUILD_FMEA_PREVIEW_TOOL_NAME,
    CandidateSimilarityRetrieve,
    ConversationFmeaPerceive,
    FmeaPreviewPlan,
    FmeaPreviewReflect,
    FmeaPreviewToolAction,
    FmeaPreviewToolDispatcher,
    build_fmea_workbook,
    decode_chat_log,
    display_fmea_rows,
    normalise_fmea_row,
    normalise_fmea_rows,
)
from src.my_splitter import FmeaDocument


class FakeEmbeddings:
    def __init__(self, vectors: list[list[float]]) -> None:
        self.vectors = vectors
        self.calls: list[dict[str, Any]] = []

    def create(self, **kwargs: Any) -> SimpleNamespace:
        self.calls.append(kwargs)
        return SimpleNamespace(
            data=[
                SimpleNamespace(index=index, embedding=vector)
                for index, vector in enumerate(self.vectors)
            ]
        )


class FakeEmbeddingClient:
    def __init__(self, vectors: list[list[float]]) -> None:
        self.embeddings = FakeEmbeddings(vectors)


class FakeKnowledgeBase:
    def __init__(
        self,
        process_code: str,
        vectors: list[list[float]],
    ) -> None:
        matrix = np.asarray(vectors, dtype="float32")
        faiss.normalize_L2(matrix)
        self.index = faiss.IndexFlatIP(matrix.shape[1])
        self.index.add(matrix)
        self.documents = [
            FmeaDocument(
                document_id=f"{process_code}-{index + 1:04d}",
                content="existing row",
                metadata={},
            )
            for index in range(len(vectors))
        ]


def _raw_row(**values: Any) -> dict[str, Any]:
    return {field: values.get(field) for field in STANDARD_FIELDS}


def test_normalise_row_has_exact_schema_and_only_derives_complete_rpn() -> None:
    row = normalise_fmea_row(
        {
            "process": "清洗",
            "potential_failure_mode": "殘留污染",
            "severity_before": "8",
            "occurrence_before": 3,
            "detection_before": "2",
            "severity_after": 4,
            "occurrence_after": 2,
            "unexpected": "ignored",
        }
    )

    assert list(row) == STANDARD_FIELDS
    assert row["rpn_before"] == 48
    assert row["rpn_after"] is None
    assert row["owner_date"] is None
    assert "unexpected" not in row


def test_normalise_rows_removes_exact_duplicate_events_and_empty_rows() -> None:
    event = {
        "process": "PVD",
        "potential_failure_mode": "鍍膜不足",
        "potential_causes": "壓力異常",
    }

    rows = normalise_fmea_rows([event, dict(event), {}, "invalid"])

    assert len(rows) == 1
    assert rows[0]["potential_failure_mode"] == "鍍膜不足"


def test_visible_rows_and_workbook_use_null_and_canonical_headers() -> None:
    rows = [normalise_fmea_row({"process": "PVD", "severity_before": 8})]

    visible = display_fmea_rows(rows)
    workbook_bytes = build_fmea_workbook(rows)

    assert visible[0]["potential_failure_mode"] == "NULL"
    assert workbook_bytes is not None
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True)
    worksheet = workbook.active
    values = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    assert list(values[0]) == STANDARD_FIELDS
    assert values[1][0] == "PVD"
    assert values[1][2] == "NULL"
    assert len(values) == 2


def test_perceive_extracts_and_normalises_candidate_rows() -> None:
    response = OpenAIChatResponse(
        content=(
            '{"candidate_rows": [{"process":"PVD",'
            '"potential_failure_mode":"鍍膜不足",'
            '"potential_causes":"腔體壓力異常",'
            '"severity_before":8,"extra":"ignored"}]}'
        ),
        model="test-model",
        input_tokens=10,
        output_tokens=12,
    )

    def runner(client: object, **kwargs: Any) -> OpenAIChatResponse:
        return response

    perceive = ConversationFmeaPerceive(
        model="test-model",
        client=object(),
        chat_runner=runner,
    )
    output = perceive(WorkflowState(user_message="PVD 腔體壓力異常造成鍍膜不足"))

    rows = output["payload"]["candidate_rows"]
    assert len(rows) == 1
    assert list(rows[0]) == STANDARD_FIELDS
    assert rows[0]["potential_causes"] == "腔體壓力異常"
    assert rows[0]["occurrence_before"] is None


def test_perceive_provider_failure_aborts_with_safe_message() -> None:
    def runner(client: object, **kwargs: Any) -> OpenAIChatResponse:
        raise RuntimeError("secret provider detail")

    perceive = ConversationFmeaPerceive(
        model="test-model",
        client=object(),
        chat_runner=runner,
    )

    with pytest.raises(WorkflowAborted, match="無法從聊天紀錄整理"):
        perceive(WorkflowState(user_message="聊天內容"))


def test_similarity_retrieve_batches_embeddings_and_rejects_equal_threshold() -> None:
    client = FakeEmbeddingClient([[1.0, 0.0], [0.0, 1.0]])
    knowledge_bases = {
        "PVD": FakeKnowledgeBase("PVD", [[1.0, 0.0]]),
        "PI": FakeKnowledgeBase("PI", [[0.7, 0.7]]),
    }
    retrieve = CandidateSimilarityRetrieve(
        knowledge_bases,  # type: ignore[arg-type]
        embedding_client=client,
        embedding_model="embedding-model",
        threshold=1.0,
    )
    state = WorkflowState(user_message="chat")
    state.entities.update(
        {
            "candidate_rows": [
                _raw_row(process="PVD", potential_failure_mode="duplicate"),
                _raw_row(process="PVD", potential_failure_mode="new"),
            ]
        }
    )

    output = retrieve(state)

    assert len(client.embeddings.calls) == 1
    assert len(client.embeddings.calls[0]["input"]) == 2
    assert len(output["payload"]["duplicate_rows"]) == 1
    duplicate = output["payload"]["duplicate_rows"][0]
    assert duplicate["matched_process"] == "PVD"
    assert duplicate["matched_document_id"] == "PVD-0001"
    assert duplicate["similarity"] == pytest.approx(1.0)
    assert len(output["payload"]["accepted_rows"]) == 1


def test_similarity_retrieve_rejects_wrong_embedding_dimension() -> None:
    retrieve = CandidateSimilarityRetrieve(
        {"PVD": FakeKnowledgeBase("PVD", [[1.0, 0.0]])},  # type: ignore[arg-type]
        embedding_client=FakeEmbeddingClient([[1.0, 0.0, 0.0]]),
        embedding_model="embedding-model",
    )
    state = WorkflowState(user_message="chat")
    state.entities.update(
        {"candidate_rows": [_raw_row(potential_failure_mode="failure")]}
    )

    with pytest.raises(ValueError, match="index 維度"):
        retrieve(state)


def test_preview_tool_dispatcher_is_allow_listed() -> None:
    dispatcher = FmeaPreviewToolDispatcher()

    with pytest.raises(ValueError, match="不允許的工具"):
        dispatcher.execute("unknown_tool", {}, rows=[])
    with pytest.raises(ValueError, match="不接受參數"):
        dispatcher.execute(
            BUILD_FMEA_PREVIEW_TOOL_NAME,
            {"path": "somewhere"},
            rows=[],
        )


def test_agentic_preview_workflow_visits_all_stages_and_returns_entities() -> None:
    response = OpenAIChatResponse(
        content=(
            '{"candidate_rows":[{"process":"PVD",'
            '"potential_failure_mode":"新失效"}]}'
        ),
        model="test-model",
    )

    def runner(client: object, **kwargs: Any) -> OpenAIChatResponse:
        return response

    perceive = ConversationFmeaPerceive(
        model="test-model", client=object(), chat_runner=runner
    )
    workflow = Workflow(
        workflow_name="preview-test",
        perceive=perceive,
        plan=FmeaPreviewPlan(),
        retrieve=CandidateSimilarityRetrieve(
            {"PVD": FakeKnowledgeBase("PVD", [[1.0, 0.0]])},  # type: ignore[arg-type]
            embedding_client=FakeEmbeddingClient([[0.0, 1.0]]),
            embedding_model="embedding-model",
            threshold=0.85,
        ),
        action=FmeaPreviewToolAction(),
        reflect=FmeaPreviewReflect(),
    )
    events: list[dict[str, Any]] = []

    result = workflow.run("聊天紀錄", event_callback=events.append)

    started = [
        event["module"]
        for event in events
        if event.get("type") == "stage" and event.get("phase") == "start"
    ]
    assert started == ["perceive", "plan", "retrieve", "action", "reflect"]
    assert result.entities["reflect_verdict"] == "pass"
    assert len(result.entities["candidate_rows"]) == 1
    assert len(result.entities["accepted_rows"]) == 1
    assert result.entities["duplicate_rows"] == []
    assert len(result.entities["preview_rows"]) == 1
    assert isinstance(result.entities["preview_xlsx_bytes"], bytes)


def test_all_duplicates_produce_no_blank_workbook() -> None:
    action = FmeaPreviewToolAction()
    state = WorkflowState(user_message="chat")
    state.entities.update({"accepted_rows": []})

    output = action(state)

    assert output["payload"]["preview_rows"] == []
    assert output["payload"]["preview_xlsx_bytes"] is None


def test_chat_log_decoder_accepts_bom_and_rejects_non_utf8() -> None:
    assert decode_chat_log(b"\xef\xbb\xbfhello") == "hello"
    with pytest.raises(ValueError, match="UTF-8"):
        decode_chat_log(b"\xff\xfe\x00")
